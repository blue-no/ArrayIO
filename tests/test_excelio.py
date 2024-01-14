import gc
import shutil

import memory_profiler
import openpyxl

from arrayio import Array, read_excel_array, set_verbosity
from tests import CELLS, DIR, VALUES

fp_in = DIR / "input.xlsx"
shname_in = "input"

fp_base = DIR / "output_base.xlsx"
fp_out1 = DIR / "output1.xlsx"
fp_out2 = DIR / "output2.xlsx"
shname_base = "base"
shname_out1 = "output"
shname_out2 = "output"
cell_out = "B3"

set_verbosity(True)


@memory_profiler.profile
def test_read_excel():
    print("\n" + test_read_excel.__name__)
    book = openpyxl.load_workbook(
        filename=fp_in, read_only=True, data_only=True, keep_links=False
    )
    array = read_excel_array(
        sheet=book[shname_in],
        range_=CELLS,
        lazy=False,
    )
    assert array.get_values() == VALUES
    gc.collect()


@memory_profiler.profile
def test_write_excel():
    print("\n" + test_write_excel.__name__)
    shutil.copyfile(fp_base, fp_out1)
    book = openpyxl.load_workbook(filename=fp_out1)
    sheet = book.copy_worksheet(book[shname_base])
    sheet.title = shname_out1

    Array(values=VALUES).write_to_excel(
        sheet=sheet,
        cell=cell_out,
    )
    book.save(filename=fp_out1)
    gc.collect()


@memory_profiler.profile
def test_excel_io_lazy():
    print("\n" + test_excel_io_lazy.__name__)
    shutil.copyfile(fp_base, fp_out2)
    book_in = openpyxl.load_workbook(
        filename=fp_in, read_only=True, data_only=True, keep_links=False
    )
    book_out = openpyxl.load_workbook(filename=fp_out2)
    sh_out = book_out.copy_worksheet(book_out[shname_base])
    sh_out.title = shname_out2

    array = read_excel_array(
        sheet=book_in[shname_in],
        range_=CELLS,
        lazy=True,
    )

    array.write_to_excel(sheet=sh_out, cell=cell_out)

    assert array.get_values() == VALUES

    array.write_to_excel(sheet=sh_out, cell=cell_out)

    book_out.save(filename=fp_out2)
    gc.collect()
